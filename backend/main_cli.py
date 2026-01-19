try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import pandas as pd
import argparse
import os
import json
import sys

# Garante que o Python encontre a pasta src
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from src.detector import PIIDetector

def main():
    parser = argparse.ArgumentParser(description="Detector PII Participa DF - CLI")
    parser.add_argument('--input', type=str, required=True, help="Arquivo de entrada")
    parser.add_argument('--output', type=str, required=True, help="Nome do arquivo de saída")
    parser.add_argument('--usar-gpu', action='store_true', help="Usar GPU para modelos NER")
    parser.add_argument('--use-llm-arbitration', action='store_true', help="Usar árbitro LLM para casos ambíguos")
    parser.add_argument('--force-llm', action='store_true', help="Forçar uso do árbitro LLM em todos os casos")
    args = parser.parse_args()

    print("🚀 Iniciando Motor Participa DF (Híbrido + LGPD)...")
    detector = PIIDetector(
        usar_gpu=args.usar_gpu,
        use_llm_arbitration=args.use_llm_arbitration
    )

    # Leitura do arquivo
    try:
        df = pd.read_excel(args.input) if args.input.endswith(('.xlsx', '.xls')) else pd.read_csv(args.input)
    except Exception as e:
        print(f"⚠️ Erro ao ler arquivo de entrada: {e}")
        return
    
    coluna = 'Texto Mascarado'
    if coluna not in df.columns:
        print(f"⚠️ Erro: Coluna '{coluna}' não encontrada no arquivo.")
        return

    print(f"🔍 Analisando {len(df)} registros...")
    
    # Processamento - O detector retorna: (bool, lista_detalhes, risco, confianca)
    def run_detect(texto):
        return detector.detect(texto, force_llm=args.force_llm)
    results = df[coluna].fillna("").apply(run_detect)
    
    # CRIANDO AS COLUNAS NA ORDEM PADRONIZADA
    # 1. Classificação | 2. Confiança | 3. Nível de Risco | 4. Identificadores
    df['Classificação'] = results.apply(lambda x: "❌ NÃO PÚBLICO" if x[0] else "✅ PÚBLICO")
    df['Confiança'] = results.apply(lambda x: f"{x[3] * 100:.1f}%" if x[3] <= 1 else f"{x[3]:.1f}%")
    df['Nível de Risco'] = results.apply(lambda x: x[2])
    df['Identificadores'] = results.apply(lambda x: str([f"{f['tipo']}: {f['valor']}" for f in x[1]]))

    output_base = args.output.replace('.xlsx', '').replace('.csv', '').replace('.json', '')

    # ORDEM PADRONIZADA DAS COLUNAS DE RESULTADO:
    # 1. ID | 2. Texto Mascarado | 3. Classificação | 4. Confiança | 5. Nível de Risco | 6. Identificadores
    
    # Reorganiza o DataFrame para ter as colunas na ordem correta
    colunas_base = [col for col in df.columns if col not in ['Classificação', 'Confiança', 'Nível de Risco', 'Identificadores']]
    colunas_resultado = ['Classificação', 'Confiança', 'Nível de Risco', 'Identificadores']
    df = df[colunas_base + colunas_resultado]

    # 1. SALVAR JSON (Mesma estrutura e ordem do CSV/XLSX)
    json_output = []
    for index, row in df.iterrows():
        # Pega o ID original se existir, senão usa o índice
        original_id = row['ID'] if 'ID' in df.columns else index
        
        json_output.append({
            "id": str(original_id),
            "texto_mascarado": row[coluna] if coluna in df.columns else "",
            "classificacao": row['Classificação'],
            "confianca": row['Confiança'],
            "nivel_risco": row['Nível de Risco'],
            "identificadores": row['Identificadores']
        })
    with open(f"{output_base}.json", 'w', encoding='utf-8') as f:
        json.dump(json_output, f, indent=4, ensure_ascii=False)

    # 2. SALVAR CSV (Segue a ordem do DataFrame)
    df.to_csv(f"{output_base}.csv", index=False, encoding='utf-8-sig')

    # 3. SALVAR EXCEL COM CORES
    try:
        from openpyxl.styles import PatternFill
        xlsx_path = f"{output_base}.xlsx"
        
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Relatório de Privacidade')
            ws = writer.sheets['Relatório de Privacidade']
            
            # Matriz de Cores Atualizada
            colors = {
                "CRÍTICO": PatternFill(start_color="FF9C0006", end_color="FF9C0006", fill_type="solid"), # Vermelho Escuro ARGB
                "CRITICO": PatternFill(start_color="FF9C0006", end_color="FF9C0006", fill_type="solid"), # Vermelho Escuro ARGB (sem acento)
                "ALTO": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),    # Vermelho Claro ARGB
                "MODERADO": PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"), # Amarelo ARGB
                "BAIXO": PatternFill(start_color="FFDEEBF7", end_color="FFDEEBF7", fill_type="solid"),    # Azul Claro ARGB
                "SEGURO": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")   # Verde ARGB
            }

            # Localiza dinamicamente as colunas para pintar, mesmo se a ordem mudar
            col_risk_idx = df.columns.get_loc('Nível de Risco') + 1
            col_class_idx = df.columns.get_loc('Classificação') + 1

            for row in range(2, len(df) + 2):
                # Pinta a célula de Nível de Risco (normalizando para garantir cor correta)
                risk_val_raw = ws.cell(row=row, column=col_risk_idx).value
                risk_val_norm = str(risk_val_raw).strip().upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Ç", "C")
                # Aplica cor usando o dicionário normalizado
                if risk_val_norm in colors:
                    ws.cell(row=row, column=col_risk_idx).fill = colors[risk_val_norm]

                # Pinta a célula de Classificação: verde para público, vermelho para não público
                class_val = str(ws.cell(row=row, column=col_class_idx).value)
                if "NÃO" in class_val:
                    ws.cell(row=row, column=col_class_idx).fill = colors["ALTO"]  # vermelho claro
                else:
                    ws.cell(row=row, column=col_class_idx).fill = colors["SEGURO"]  # verde

        print(f"✅ Sucesso! Relatórios gerados em: {os.path.dirname(os.path.abspath(xlsx_path))}")
        print(f"📊 Ordem das colunas: Classificação -> Confiança -> Nível de Risco -> Identificadores")
        
    except Exception as e:
        print(f"⚠️ Erro na formatação Excel: {e}")
        df.to_excel(f"{output_base}.xlsx", index=False)

if __name__ == "__main__":
    main()