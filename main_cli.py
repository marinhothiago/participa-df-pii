import pandas as pd
import argparse
import os
import json
import sys

# Garante que o Python encontre a pasta src
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from src.detector import PIIDetector

def main():
    parser = argparse.ArgumentParser(description="Detector PII Participa DF - CLI")
    parser.add_argument('--input', type=str, required=True, help="Arquivo de entrada")
    parser.add_argument('--output', type=str, required=True, help="Nome do arquivo de saída")
    args = parser.parse_args()

    print("🚀 Iniciando Motor Participa DF (Híbrido + LGPD)...")
    detector = PIIDetector()

    # Leitura do arquivo
    try:
        df = pd.read_excel(args.input) if args.input.endswith(('.xlsx', '.xls')) else pd.read_csv(args.input)
    except Exception as e:
        print(f"⚠️ Erro ao ler arquivo de entrada: {e}")
        return
    
    coluna = 'Texto Mascarado'
    if coluna not in df.columns:
        print(f"⚠️ Erro: Coluna '{coluna}' não encontrada no arquivo.")
        return

    print(f"🔍 Analisando {len(df)} registros...")
    
    # Processamento - O detector retorna: (bool, lista_detalhes, risco, confianca)
    results = df[coluna].fillna("").apply(detector.detect)
    
    # CRIANDO AS COLUNAS NA ORDEM PADRONIZADA
    # 1. Classificação | 2. Confiança | 3. Nível de Risco | 4. Identificadores
    df['Classificação'] = results.apply(lambda x: "❌ NÃO PÚBLICO" if x[0] else "✅ PÚBLICO")
    df['Confiança'] = results.apply(lambda x: f"{x[3] * 100:.1f}%" if x[3] <= 1 else f"{x[3]:.1f}%")
    df['Nível de Risco'] = results.apply(lambda x: x[2])
    df['Identificadores'] = results.apply(lambda x: str([f"{f['tipo']}: {f['valor']}" for f in x[1]]))

    output_base = args.output.replace('.xlsx', '').replace('.csv', '').replace('.json', '')

    # 1. SALVAR JSON (Preservando o ID original da planilha)
    json_output = []
    for index, r in enumerate(results):
        # Tenta pegar o valor da coluna 'ID' se ela existir, senão usa o índice
        original_id = df['ID'].iloc[index] if 'ID' in df.columns else index
        
        json_output.append({
            "id": str(original_id), # Garante o ID no JSON
            "classificacao": "NAO_PUBLICO" if r[0] else "PUBLICO",
            "confianca": f"{r[3] * 100:.1f}%",
            "risco": r[2],
            "detalhes": r[1]
        })
    with open(f"{output_base}.json", 'w', encoding='utf-8') as f:
        json.dump(json_output, f, indent=4, ensure_ascii=False)

    # 2. SALVAR CSV (Segue a ordem do DataFrame)
    df.to_csv(f"{output_base}.csv", index=False, encoding='utf-8-sig')

    # 3. SALVAR EXCEL COM CORES
    try:
        from openpyxl.styles import PatternFill
        xlsx_path = f"{output_base}.xlsx"
        
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Relatório de Privacidade')
            ws = writer.sheets['Relatório de Privacidade']
            
            # Matriz de Cores Atualizada
            colors = {
                "CRÍTICO": PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid"), # Vermelho Escuro
                "ALTO": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),    # Vermelho Claro
                "MODERADO": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), # Amarelo
                "BAIXO": PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid"),    # Azul Claro
                "SEGURO": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Verde
            }

            # Localiza dinamicamente as colunas para pintar, mesmo se a ordem mudar
            col_risk_idx = df.columns.get_loc('Nível de Risco') + 1
            col_class_idx = df.columns.get_loc('Classificação') + 1

            for row in range(2, len(df) + 2):
                # Pinta a célula de Nível de Risco
                risk_val = ws.cell(row=row, column=col_risk_idx).value
                if risk_val in colors:
                    ws.cell(row=row, column=col_risk_idx).fill = colors[risk_val]
                
                # Pinta a célula de Classificação
                class_val = str(ws.cell(row=row, column=col_class_idx).value)
                color_key = "ALTO" if "NÃO" in class_val else "SEGURO"
                ws.cell(row=row, column=col_class_idx).fill = colors[color_key]

        print(f"✅ Sucesso! Relatórios gerados em: {os.path.dirname(os.path.abspath(xlsx_path))}")
        print(f"📊 Ordem das colunas: Classificação -> Confiança -> Nível de Risco -> Identificadores")
        
    except Exception as e:
        print(f"⚠️ Erro na formatação Excel: {e}")
        df.to_excel(f"{output_base}.xlsx", index=False)

if __name__ == "__main__":
    main()